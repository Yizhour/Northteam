import copy
from pathlib import Path

import pandas as pd
from django.apps import apps
from django.conf import settings
from django.db import DatabaseError, transaction
from openpyxl import load_workbook
from openpyxl.styles.colors import COLOR_INDEX

from .config import CONFIG_FILE, TABLE_CACHE_FILE, ensure_directories


DATE_COLUMNS = [
    "预计申报日期/发行日期/存续期归档起算日",
    "归档流程发起截止日",
    "协会报送截止日",
]

DISPLAY_COLUMNS = [
    "序号",
    "项目名称",
    "项目编号",
    "项目负责人",
    "底稿报送阶段",
    "预计申报日期/发行日期/存续期归档起算日",
    "归档流程发起截止日",
    "距归档流程发起截止日",
    "是否已完成归档",
    "协会报送截止日",
    "距协会报送截止日剩余工作日",
    "是否已报送协会",
]

DEFAULT_CONFIG = {
    "sender_email": "",
    "auth_code": "",
    "receiver_list": [],
    "excel_path": str(TABLE_CACHE_FILE),
    "weekly_enabled": True,
    "weekly_time": "09:00",
    "schedule_day": "Monday",
    "daily_enabled": False,
    "daily_time": "09:00",
    "daily_email_receiver": "",
    "send_interval_seconds": 30,
    "smtp_preferred_mode": "auto",
    "smtp_timeout_seconds": 8,
    "header_row_index": 0,
    "ui_original_path": "",
    "ui_header_row_index": 0,
    "owner_column": "项目负责人",
    "owner_name": "潘学超",
    "date_columns": DATE_COLUMNS,
    "display_columns": DISPLAY_COLUMNS,
    "archive_deadline_column": "归档流程发起截止日",
    "association_deadline_column": "协会报送截止日",
    "association_workday_threshold": 10,
    "column_widths": {},
    "table_widths": {},
    "column_colors": {
        "预计申报日期/发行日期/存续期归档起算日": "#3498db",
        "归档流程发起截止日": "#e74c3c",
        "协会报送截止日": "#f39c12",
    },
    "column_sms_texts": {
        "预计申报日期/发行日期/存续期归档起算日": "预计申报/发行/归档起算",
        "归档流程发起截止日": "归档流程发起截止",
        "协会报送截止日": "协会报送截止",
    },
    "default_column_mappings": {
        "date_columns": DATE_COLUMNS,
        "display_columns": DISPLAY_COLUMNS,
    },
    "email_subject": "【底稿报送提醒】关键节点提醒",
    "email_intro": "您好，底稿报送提醒事项如下：",
    "daily_msg_template": "“{项目名称}” {短信文本}",
    "daily_msg_intro": "您好，今日有{n}项底稿报送事项需要处理：",
}

STORE_CONFIG = "config"
STORE_META = "table_meta"


def _models():
    if not settings.configured or not apps.ready:
        return None, None
    try:
        store_model = apps.get_model("manuscriptreminder", "ManuscriptReminderStore")
        row_model = apps.get_model("manuscriptreminder", "ManuscriptReminderTableRow")
        return store_model, row_model
    except LookupError:
        return None, None


def _load_store(key, default):
    store_model, _ = _models()
    if store_model is None:
        return copy.deepcopy(default)
    try:
        item, _ = store_model.objects.get_or_create(key=key, defaults={"value": copy.deepcopy(default)})
        return copy.deepcopy(item.value)
    except DatabaseError:
        return copy.deepcopy(default)


def _save_store(key, value):
    store_model, _ = _models()
    if store_model is None:
        return value
    item, _ = store_model.objects.get_or_create(key=key, defaults={"value": value})
    item.value = value
    item.save(update_fields=["value", "updated_at"])
    return value


def _cell_to_json(value):
    if pd.isna(value):
        return ""
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            pass
    return str(value).strip()


def _normalized_headers(values):
    return [str(value or "").strip().replace("\n", "") for value in values]


def _color_to_rgb(color):
    if not color:
        return ""
    if color.type == "rgb" and color.rgb:
        return str(color.rgb).upper()
    if color.type == "indexed" and color.indexed is not None:
        try:
            value = COLOR_INDEX[int(color.indexed)]
            return str(value).upper() if value else f"INDEXED:{color.indexed}"
        except (IndexError, TypeError, ValueError):
            return f"INDEXED:{color.indexed}"
    if color.type == "theme":
        return f"THEME:{color.theme}"
    return ""


def _classify_alert_fill(cell):
    fill = cell.fill
    if not fill or not fill.fill_type:
        return ""
    code = _color_to_rgb(fill.fgColor)
    compact = code[-6:] if code else ""
    if compact in {"FFFF00", "FFFF99"}:
        return "yellow"
    if compact in {"FF0000", "C00000", "FF6666"}:
        return "red"
    return ""


def _extract_xlsx(path, header=0):
    workbook = load_workbook(path, read_only=False, data_only=True)
    worksheet = workbook.active
    header_row = int(header) + 1
    columns = _normalized_headers([worksheet.cell(header_row, col).value for col in range(1, worksheet.max_column + 1)])
    rows = []
    styles = []
    for excel_row in range(header_row + 1, worksheet.max_row + 1):
        values = [worksheet.cell(excel_row, col).value for col in range(1, worksheet.max_column + 1)]
        if not any(str(value or "").strip() for value in values):
            continue
        row_data = {columns[index]: _cell_to_json(value) for index, value in enumerate(values) if columns[index]}
        alert_colors = []
        for col in range(1, worksheet.max_column + 1):
            alert = _classify_alert_fill(worksheet.cell(excel_row, col))
            if alert:
                alert_colors.append(alert)
        fill_alert = "red" if "red" in alert_colors else ("yellow" if "yellow" in alert_colors else "")
        rows.append(row_data)
        styles.append({"excel_row": excel_row, "fill_alert": fill_alert})
    return columns, rows, styles


def _extract_plain_table(path, header=0):
    path = Path(path)
    if path.suffix.lower() == ".csv":
        last_error = None
        for encoding in ["utf-8-sig", "gb18030", "gbk"]:
            try:
                df = pd.read_csv(path, header=header, encoding=encoding, sep=None, engine="python")
                break
            except UnicodeDecodeError as exc:
                last_error = exc
        else:
            raise last_error
    else:
        df = pd.read_excel(path, header=header)
    df.columns = _normalized_headers(df.columns)
    rows = [
        {column: _cell_to_json(row[column]) for column in df.columns}
        for _, row in df.iterrows()
    ]
    return list(df.columns), rows, [{"excel_row": index + int(header) + 2, "fill_alert": ""} for index in range(len(rows))]


def _save_table(columns, rows, styles, source_path=""):
    _, row_model = _models()
    if row_model is None:
        raise RuntimeError("数据库未就绪，无法保存底稿报送数据。")
    records = [
        row_model(row_index=index, data={column: row.get(column, "") for column in columns}, style=styles[index] if index < len(styles) else {})
        for index, row in enumerate(rows)
    ]
    with transaction.atomic():
        row_model.objects.all().delete()
        if records:
            row_model.objects.bulk_create(records, batch_size=500)
        _save_store(
            STORE_META,
            {
                "columns": columns,
                "total_rows": len(records),
                "source_path": str(source_path),
            },
        )


def has_table():
    _, row_model = _models()
    if row_model is None:
        return False
    try:
        return row_model.objects.exists()
    except DatabaseError:
        return False


def load_config():
    ensure_directories()
    data = _load_store(STORE_CONFIG, DEFAULT_CONFIG)
    for key, value in DEFAULT_CONFIG.items():
        data.setdefault(key, copy.deepcopy(value))
    data["excel_path"] = str(TABLE_CACHE_FILE)
    return data


def save_config(data):
    merged = load_config()
    merged.update(data)
    merged["excel_path"] = str(TABLE_CACHE_FILE)
    return _save_store(STORE_CONFIG, merged)


def public_config(data=None):
    data = copy.deepcopy(data or load_config())
    if data.get("auth_code"):
        data["auth_code"] = ""
        data["auth_code_set"] = True
    else:
        data["auth_code_set"] = False
    return data


def save_table_from_upload(source_path, header=0, source_name=""):
    path = Path(source_path)
    if path.suffix.lower() == ".xlsx":
        columns, rows, styles = _extract_xlsx(path, header=header)
    else:
        columns, rows, styles = _extract_plain_table(path, header=header)
    source_label = source_name or path.name
    _save_table(columns, rows, styles, source_path=source_label)
    config = load_config()
    config["ui_original_path"] = source_label
    config["ui_header_row_index"] = int(header)
    save_config(config)
    return columns, rows


def load_table_rows(limit=None):
    _, row_model = _models()
    if row_model is None:
        return []
    try:
        query = row_model.objects.order_by("row_index")
        if limit is not None:
            query = query[:limit]
        return [{"data": item.data, "style": item.style or {}, "row_index": item.row_index} for item in query]
    except DatabaseError:
        return []


def table_preview(limit=50):
    meta = _load_store(STORE_META, {"columns": [], "total_rows": 0, "source_path": ""})
    columns = [str(column) for column in meta.get("columns", [])]
    rows = load_table_rows(limit=limit)
    return {
        "columns": columns,
        "rows": [item["data"] for item in rows],
        "total_rows": int(meta.get("total_rows") or 0),
        "path": meta.get("source_path") or "",
    }
